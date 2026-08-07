#!/usr/bin/env python3
"""One-shot inventory of DB DUMP folder for N-087 discovery (read-only)."""
from __future__ import annotations

import csv
import io
import json
import pickle
import re
import sqlite3
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

DUMP = Path(r"y:\Digital Stealth Care\Projects\DB DUMP")
DATA = Path(r"y:\Digital Stealth Care\Projects\DSC-HUB\homeassistant\data")
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from catalog_common import name_norm  # noqa: E402


def zip_inventory() -> None:
    print("=== ZIP CONTENTS ===")
    for zpath in sorted(DUMP.glob("*.zip")):
        print(f"\n{zpath.name} ({zpath.stat().st_size:,} bytes)")
        with zipfile.ZipFile(zpath) as z:
            infos = z.infolist()
            print(f"  members={len(infos)}")
            by_ext: dict[str, int] = defaultdict(int)
            dataish = []
            for i in infos:
                ext = Path(i.filename).suffix.lower() or "(none)"
                by_ext[ext] += 1
                if (
                    ext
                    in {
                        ".csv",
                        ".json",
                        ".jsonl",
                        ".tsv",
                        ".xlsx",
                        ".xls",
                        ".sql",
                        ".db",
                        ".parquet",
                        ".pkl",
                        ".p",
                        ".txt",
                        ".arff",
                    }
                    and not i.is_dir()
                ):
                    dataish.append((i.filename, i.file_size))
            print("  exts:", dict(sorted(by_ext.items(), key=lambda x: -x[1])[:12]))
            for fn, sz in sorted(dataish, key=lambda x: -x[1])[:30]:
                print(f"  DATA {sz:>10,}  {fn}")
            if len(dataish) > 30:
                print(f"  ... +{len(dataish) - 30} more dataish")


def list_data_files() -> list[tuple[int, str, str, Path]]:
    skip = {".git", "node_modules", "__pycache__", "dist", "build", ".next"}
    exts = {
        ".csv",
        ".json",
        ".jsonl",
        ".tsv",
        ".xlsx",
        ".xls",
        ".xlsm",
        ".sql",
        ".db",
        ".arff",
        ".parquet",
        ".pkl",
        ".p",
    }
    rows: list[tuple[int, str, str, Path]] = []
    for p in DUMP.rglob("*"):
        if not p.is_file():
            continue
        if any(part in skip for part in p.parts):
            continue
        if p.suffix.lower() not in exts:
            continue
        sz = p.stat().st_size
        if sz < 8_000 and p.suffix.lower() not in {".parquet", ".db", ".sql", ".arff"}:
            continue
        rel = str(p.relative_to(DUMP))
        rows.append((sz, rel, p.suffix.lower(), p))
    rows.sort(reverse=True)
    return rows


def peek_csv_header(path: Path) -> tuple[list[str], int]:
    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        r = csv.DictReader(f)
        cols = list(r.fieldnames or [])
        n = 0
        for _ in r:
            n += 1
        return cols, n


def load_dump_name_norms(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        doc = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return set()
    items = doc.get("items") or []
    out = set()
    for it in items:
        if not isinstance(it, dict):
            continue
        nn = it.get("name_norm")
        if nn:
            out.add(str(nn))
            continue
        name = it.get("name") or it.get("Strain") or it.get("strain_name")
        if name:
            out.add(name_norm(str(name)))
    return out


def corpus_name_norms_sample() -> set[str]:
    """Union of key existing dumps for overlap estimates (avoid locked sqlite)."""
    names: set[str] = set()
    for p in [
        DATA / "dsc_strains_seedcity_local.json",
        DATA / "dsc_strains_seedcity.json",
        DATA / "dsc_strains_leafly_flat.json",
        DATA / "dsc_strains_leafly_features.json",
        DATA / "dsc_strains_leafly_kaggle.json",
        DATA / "dsc_strains_leafly_github.json",
        DATA / "dsc_strains_kushy.json",
        DATA / "dsc_strains_openthc.json",
        DATA / "dsc_strains_wikileaf.json",
        DATA / "dsc_strains_pickle_archive.json",
        DATA / "dsc_strains_project_lists.json",
        DATA / "dsc_strains_mj_simple.json",
        DATA / "dsc_strains_lynch_figshare.json",
        DATA / "dsc_lab_replication_wa.json",
    ]:
        names |= load_dump_name_norms(p)
    return names


def estimate_csv_overlap(path: Path, name_cols: list[str], corpus: set[str], *, slug: bool = False) -> dict:
    cols, n = peek_csv_header(path)
    uniq: set[str] = set()
    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        for r in csv.DictReader(f):
            name = ""
            for c in name_cols:
                if c in r and r[c]:
                    name = str(r[c]).strip()
                    break
            if not name:
                continue
            if slug or ("-" in name and " " not in name and len(name) < 60):
                # treat as slug-ish
                nn = name_norm(name.replace("-", " ").replace("_", " "))
            else:
                nn = name_norm(name)
            if nn:
                uniq.add(nn)
    new = uniq - corpus
    return {
        "rows": n,
        "cols": cols,
        "unique_names": len(uniq),
        "overlap": len(uniq & corpus),
        "new_est": len(new),
        "sample_new": sorted(list(new))[:8],
    }


def pickle_viability() -> None:
    print("\n=== PICKLE VIABILITY (numpy) ===")
    pdir = DUMP / "archive (3)" / "cannabis 2" / "Strain data" / "strains"
    if not pdir.exists():
        # try zip
        zpath = DUMP / "archive (3).zip"
        if not zpath.exists():
            print("missing pickle archive")
            return
        with zipfile.ZipFile(zpath) as z:
            names = [n for n in z.namelist() if n.endswith(".p")]
            print(f"zip pickles={len(names)}")
            sample = names[0]
            raw = z.read(sample)
    else:
        files = list(pdir.glob("*.p"))
        print(f"unzipped pickles={len(files)}")
        sample = files[0]
        raw = sample.read_bytes()
        sample = sample.name

    try:
        import numpy as np  # noqa: F401

        print(f"numpy ok: {np.__version__}")
        obj = pickle.loads(raw)
        print(f"sample={sample} type={type(obj)}")
        if isinstance(obj, dict):
            print(f"keys={list(obj.keys())}")
            for k, v in obj.items():
                t = type(v).__name__
                if hasattr(v, "shape"):
                    print(f"  {k}: ndarray shape={getattr(v,'shape',None)} dtype={getattr(v,'dtype',None)}")
                elif isinstance(v, list):
                    print(f"  {k}: list len={len(v)} first_type={type(v[0]).__name__ if v else None}")
                    if v and isinstance(v[0], dict):
                        print(f"    first keys={list(v[0].keys())[:12]}")
                else:
                    print(f"  {k}: {t} {str(v)[:120]}")
        print("viability: FULL with numpy (categorias ndarray + data_strain review list)")
    except Exception as exc:
        print(f"numpy pickle fail: {exc}")
        # best-effort already known from importer
        print("viability: PARTIAL without numpy (Dummy unpickler strips arrays)")


def parquet_schema() -> None:
    print("\n=== PARQUET (no blob load) ===")
    path = DUMP / "train-00000-of-00002.parquet"
    import pyarrow.parquet as pq

    pf = pq.ParquetFile(path)
    md = pf.metadata
    print(f"rows={md.num_rows} row_groups={md.num_row_groups} cols={md.num_columns}")
    for i in range(md.num_columns):
        col = md.schema.column(i)
        print(f"  col[{i}] name={col.name} physical={col.physical_type} logical={col.logical_type}")
    # only read label column
    table = pq.read_table(path, columns=["label"])
    labels = table.column("label").to_pylist()
    print(f"label sample={labels[:20]} unique={sorted(set(labels))}")
    print("NOTE: image column is struct<bytes,path> — DO NOT load for catalog")
    print("residual_value: NONE for strain catalog (vision train shard)")


def xlsx_peek() -> None:
    print("\n=== XLSX Cannabis StrainsFeatures.xlsx ===")
    path = DUMP / "Cannabis StrainsFeatures.xlsx"
    if not path.exists():
        print("missing")
        return
    try:
        import openpyxl
    except ImportError:
        try:
            import pandas as pd

            xl = pd.ExcelFile(path)
            print(f"sheets={xl.sheet_names}")
            for sn in xl.sheet_names[:3]:
                df = pd.read_excel(path, sheet_name=sn, nrows=2)
                print(f"  {sn}: cols={list(df.columns)} shape_hint_rows>=2")
            # full count cheaply via openpyxl? use pandas
            df = pd.read_excel(path, sheet_name=0)
            print(f"sheet0 rows={len(df)} cols={list(df.columns)}")
            return
        except Exception as exc:
            print(f"xlsx fail: {exc}")
            return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"sheets={wb.sheetnames}")
    for sn in wb.sheetnames[:3]:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        n = 0
        for _ in rows:
            n += 1
        print(f"  {sn}: cols={header} rows={n}")


def session_db() -> None:
    print("\n=== session.db ===")
    path = DUMP / "session.db"
    con = sqlite3.connect(str(path))
    try:
        for (t,) in con.execute("SELECT name FROM sqlite_master WHERE type='table'"):
            n = con.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
            cols = [r[1] for r in con.execute(f'PRAGMA table_info("{t}")')]
            print(f"  {t}: {n} cols={cols}")
            if t == "events" and n:
                sample = con.execute(f'SELECT event_data FROM "{t}" LIMIT 1').fetchone()
                if sample:
                    print(f"    event_data sample={str(sample[0])[:200]}")
    finally:
        con.close()
    print("residual: NONE (app session junk)")


def cannia_keys() -> None:
    print("\n=== cannia strains.json ===")
    path = DUMP / "cannia-master" / "src" / "data" / "strains.json"
    if not path.exists():
        path = DUMP / "cannia-master (1)" / "src" / "data" / "strains.json"
    doc = json.loads(path.read_text(encoding="utf-8"))
    print(f"keys={len(doc)}")
    first = next(iter(doc.values()))
    print(f"value_type={type(first).__name__} sample_keys={list(first.keys())[:20] if isinstance(first, dict) else first}")


def main() -> None:
    zip_inventory()
    rows = list_data_files()
    print(f"\n=== DATA FILES >=8KB count={len(rows)} ===")
    by = defaultdict(lambda: [0, 0])
    for sz, rel, ext, _ in rows:
        top = rel.replace("/", "\\").split("\\")[0]
        by[top][0] += 1
        by[top][1] += sz
    print("\n=== BY TOP FOLDER ===")
    for top, (n, b) in sorted(by.items(), key=lambda x: -x[1][1]):
        print(f"{b / 1e6:8.1f} MB  n={n:4}  {top}")

    print("\n=== TOP 60 DATA FILES ===")
    for sz, rel, ext, _ in rows[:60]:
        print(f"{sz:>12,}  {ext:8}  {rel}")

    print("\nLoading corpus name_norms from dumps (sqlite busy — avoid)...")
    corpus = corpus_name_norms_sample()
    print(f"corpus_union_name_norms≈{len(corpus)}")

    targets = [
        (
            DUMP / "cannabis-intelligence-database-main" / "data" / "Cannabis_Intelligence_Database_15768_Strains_Final.csv",
            ["strain_name"],
            False,
            "intelligence_db",
        ),
        (
            DUMP / "phytochemical-diversity-cannabis-main" / "data" / "strain_info_pub_20210915.csv",
            ["strain_slug"],
            True,
            "phyto_strain_info",
        ),
        (
            DUMP / "cannabis_analysis-main" / "Needed CSV files" / "kushy_strains_crosses_with_nulls.csv",
            ["name"],
            False,
            "kushy_local",
        ),
        (
            DUMP / "cannabis_analysis-main" / "Needed CSV files" / "Read" / "strain_medical_effects.csv",
            ["name"],
            False,
            "medical_effects",
        ),
        (
            DUMP / "cannabis-strains-final.csv",
            ["strain_name"],
            False,
            "seedcity_top",
        ),
        (
            DUMP / "flattened_strains.csv",
            ["name"],
            False,
            "leafly_flat_csv",
        ),
        (
            DUMP / "Replication_Data.csv",
            ["test_strain", "leafly_strain"],
            False,
            "replication",
        ),
        (
            DUMP / "archive" / "cannabis.csv",
            ["Strain"],
            False,
            "leafly_features_dup",
        ),
        (
            DUMP / "Cannabis-Strains-main" / "cannabis-strains.csv",
            ["strain_name"],
            False,
            "seedcity_repo_dup",
        ),
        (
            DUMP / "Strain project-20260807T130449Z-1-001" / "Strain project" / "MJ_strains_simple.csv",
            ["strain"],
            False,
            "mj_simple",
        ),
    ]

    print("\n=== OVERLAP ESTIMATES ===")
    for path, cols, slug, label in targets:
        if not path.exists():
            print(f"{label}: MISSING {path}")
            continue
        est = estimate_csv_overlap(path, cols, corpus, slug=slug)
        print(
            f"{label}: rows={est['rows']} unique={est['unique_names']} "
            f"overlap≈{est['overlap']} new≈{est['new_est']} "
            f"key_cols={est['cols'][:12]}"
        )
        if est["sample_new"]:
            print(f"  sample_new={est['sample_new']}")

    # cannia + north atlantic
    for label, path, key in [
        ("cannia", DUMP / "cannia-master" / "src" / "data" / "strains.json", None),
        (
            "north_atlantic",
            DUMP
            / "cannabis-intelligence-database-main"
            / "[TRASH] data scripts"
            / "scripts"
            / "North Atlantic Seed Co"
            / "Files"
            / "north_atlantic_strains_comprehensive.json",
            "strain_name",
        ),
        ("medcabinet_json", DUMP / "med-cabinet-main" / "med-cabinet" / "static" / "data" / "cannabis.json", "Strain"),
    ]:
        if not path.exists():
            print(f"{label}: MISSING")
            continue
        doc = json.loads(path.read_text(encoding="utf-8"))
        uniq = set()
        if isinstance(doc, dict) and key is None:
            for k in doc.keys():
                uniq.add(name_norm(k))
        elif isinstance(doc, list):
            for it in doc:
                if isinstance(it, dict) and key:
                    uniq.add(name_norm(str(it.get(key) or "").replace("-", " ")))
        print(
            f"{label}: unique={len(uniq)} overlap≈{len(uniq & corpus)} new≈{len(uniq - corpus)}"
        )

    # phyto lab: unique strain_slugs that are non-empty
    phyto = DUMP / "phytochemical-diversity-cannabis-main" / "data" / "preproc_lab_data_pub_20220218.csv"
    if phyto.exists():
        slugs = set()
        empty = 0
        n = 0
        with phyto.open(encoding="utf-8", errors="replace", newline="") as f:
            for r in csv.DictReader(f):
                n += 1
                s = (r.get("strain_slug") or "").strip()
                if not s:
                    empty += 1
                    continue
                slugs.add(name_norm(s.replace("-", " ")))
        print(
            f"phyto_lab: rows={n} empty_slug={empty} unique_slugs={len(slugs)} "
            f"overlap≈{len(slugs & corpus)} new≈{len(slugs - corpus)}"
        )

    parquet_schema()
    pickle_viability()
    xlsx_peek()
    session_db()
    cannia_keys()

    # dump status for importer outputs
    print("\n=== IMPORTER OUTPUT STATUS ===")
    for name in [
        "dsc_strains_seedcity_local.json",
        "dsc_strains_leafly_flat.json",
        "dsc_strains_leafly_features.json",
        "dsc_strains_leafly_kaggle.json",
        "dsc_lab_replication.json",
        "dsc_lab_replication_wa.json",
        "dsc_strains_pickle_archive.json",
        "dsc_strains_project_lists.json",
        "dsc_strains_mj_simple.json",
        "dsc_strains_parquet_train.json",
        "dsc_strains_kushy.json",
        "dsc_strains_northatlantic.json",
    ]:
        p = DATA / name
        if not p.exists():
            print(f"{name}: MISSING")
            continue
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            print(
                f"{name}: count={d.get('count')} items={len(d.get('items') or [])} "
                f"source={d.get('source')} truncated={d.get('truncated')} note={str(d.get('note') or '')[:80]}"
            )
        except Exception as exc:
            print(f"{name}: PARSE_FAIL {exc}")


if __name__ == "__main__":
    main()
